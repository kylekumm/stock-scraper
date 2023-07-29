import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

def read_tickers():
    workbook = load_workbook(filename="Ticker Names.xlsx")
    sheet = workbook.active
    
    data = []
    for i in range(1,sheet.max_row + 1):
        data.append(sheet.cell(row=i,column=1).value)

    return data

def write_data(stockData):
    workbook = load_workbook(filename="Stock Data.xlsx")
    sheet = workbook.active
    
    currentRow = sheet.max_row + 1
    sheet["A"+str(currentRow)] = stockData.get_date()
    sheet["B"+str(currentRow)] = stockData.shortName
    sheet["C"+str(currentRow)] = stockData.previousClose
    sheet["D"+str(currentRow)] = stockData.todayOpen
    sheet["E"+str(currentRow)] = stockData.todayClose
    sheet["F"+str(currentRow)] = stockData.get_percentageGain()

    workbook.save(filename="Stock Data.xlsx")

def save_file():
    filename = "Ticker Names.xlsx"

    workbook = Workbook()
    sheet = workbook.active

    workbook.save(filename=filename)